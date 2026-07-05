import io
from collections import defaultdict

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.views import View
from django.shortcuts import render

from dcim.models import Site, Rack, Device, DeviceRole
from tenancy.models import Tenant

try:
    from netbox.plugins import get_plugin_config
    ALERT_THRESHOLD = get_plugin_config('netbox_rack_report', 'alert_threshold')
except Exception:
    ALERT_THRESHOLD = 80


def compute_utilization():
    """
    Computes rack utilization by Site, Location, Device Role and Tenant.
    Returns a list of dicts with the data for the report.
    Grouping key is (site_id, location_id) so each Location within a Site
    gets its own row (racks with no Location are grouped under "No location").
    """
    group_total_u = defaultdict(float)
    group_rack_count = defaultdict(int)
    group_meta = {}

    for rack in Rack.objects.select_related('site', 'location').all():
        loc_id = rack.location_id or 0
        key = (rack.site_id, loc_id)
        group_total_u[key] += float(rack.u_height or 0)
        group_rack_count[key] += 1
        group_meta[key] = (rack.site, rack.location)

    group_used_u = defaultdict(float)
    group_device_count = defaultdict(int)
    group_role = defaultdict(lambda: defaultdict(lambda: {'count': 0, 'u': 0.0}))
    group_tenant = defaultdict(lambda: defaultdict(lambda: {'count': 0, 'u': 0.0}))

    qs = Device.objects.select_related(
        'site', 'rack', 'rack__location', 'device_type', 'role', 'tenant'
    ).exclude(rack__isnull=True)

    for device in qs:
        loc_id = device.rack.location_id or 0
        key = (device.site_id, loc_id)
        u_height = float(device.device_type.u_height if device.device_type else 0)
        role_name = device.role.name if device.role else 'No role'
        tenant_name = device.tenant.name if device.tenant else 'No tenant'

        group_device_count[key] += 1
        group_used_u[key] += u_height or 0

        group_role[key][role_name]['count'] += 1
        group_role[key][role_name]['u'] += u_height or 0

        group_tenant[key][tenant_name]['count'] += 1
        group_tenant[key][tenant_name]['u'] += u_height or 0

    # Build result, sorted by Site name then Location name
    def sort_key(key):
        site, location = group_meta[key]
        return (
            site.name if site else '',
            location.name if location else '',
        )

    results = []
    for key in sorted(group_total_u.keys(), key=sort_key):
        site, location = group_meta[key]

        total_u = group_total_u[key]
        used_u = group_used_u.get(key, 0)
        free_u = total_u - used_u
        pct = round((used_u / total_u * 100), 1) if total_u else 0
        alert = pct >= ALERT_THRESHOLD

        results.append({
            'site': site,
            'location': location,
            'rack_count': group_rack_count[key],
            'device_count': group_device_count.get(key, 0),
            'total_u': total_u,
            'used_u': used_u,
            'free_u': free_u,
            'pct': pct,
            'alert': alert,
            'roles': dict(sorted(group_role[key].items())),
            'tenants': dict(sorted(group_tenant[key].items())),
        })

    return results


class RackReportView(LoginRequiredMixin, View):
    template_name = 'netbox_rack_report/rack_report.html'

    def get(self, request):
        results = compute_utilization()
        total_u = sum(r['total_u'] for r in results)
        used_u = sum(r['used_u'] for r in results)
        alerts = [r for r in results if r['alert']]

        return render(request, self.template_name, {
            'results': results,
            'total_u': total_u,
            'used_u': used_u,
            'free_u': total_u - used_u,
            'total_pct': round(used_u / total_u * 100, 1) if total_u else 0,
            'total_devices': sum(r['device_count'] for r in results),
            'total_racks': sum(r['rack_count'] for r in results),
            'alerts': alerts,
            'alert_threshold': ALERT_THRESHOLD,
        })


class RackReportExportView(LoginRequiredMixin, View):

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl is not installed. Run: pip install openpyxl', status=500)

        results = compute_utilization()

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1F4E5F')
        warn_fill = PatternFill('solid', fgColor='FFECB3')
        danger_fill = PatternFill('solid', fgColor='FFCDD2')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(top=thin, bottom=thin, left=thin, right=thin)
        center = Alignment(horizontal='center', vertical='center')

        def style_header(ws, row, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

        def autosize(ws):
            for col in ws.columns:
                length = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 3, 12), 45)

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = 'Summary by Site'
        headers1 = ['Site', 'Location', 'Racks', 'Devices', 'Total U', 'Used U', 'Free U', '% Utilization', 'Alert']
        ws1.append(headers1)
        style_header(ws1, 1, len(headers1))

        for i, r in enumerate(results, start=2):
            ws1.append([
                r['site'].name,
                r['location'].name if r['location'] else 'No location',
                r['rack_count'],
                r['device_count'],
                r['total_u'],
                r['used_u'],
                r['free_u'],
                r['pct'],
                f"⚠️ > {ALERT_THRESHOLD}%" if r['alert'] else '✓ OK',
            ])
            if r['alert']:
                for c in range(1, len(headers1) + 1):
                    ws1.cell(row=i, column=c).fill = danger_fill
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=i, column=c).border = border
        autosize(ws1)
        ws1.freeze_panes = 'A2'

        # Sheet 2: By Role
        ws2 = wb.create_sheet('By Device Role')
        headers2 = ['Site', 'Location', 'Device Role', 'Devices', 'Used U']
        ws2.append(headers2)
        style_header(ws2, 1, len(headers2))
        row_idx = 2
        for r in results:
            loc_name = r['location'].name if r['location'] else 'No location'
            for role, vals in r['roles'].items():
                ws2.append([r['site'].name, loc_name, role, vals['count'], vals['u']])
                for c in range(1, len(headers2) + 1):
                    ws2.cell(row=row_idx, column=c).border = border
                row_idx += 1
        autosize(ws2)
        ws2.freeze_panes = 'A2'

        # Sheet 3: By Tenant
        ws3 = wb.create_sheet('By Tenant')
        headers3 = ['Site', 'Location', 'Tenant', 'Devices', 'Used U']
        ws3.append(headers3)
        style_header(ws3, 1, len(headers3))
        row_idx = 2
        for r in results:
            loc_name = r['location'].name if r['location'] else 'No location'
            for tenant, vals in r['tenants'].items():
                ws3.append([r['site'].name, loc_name, tenant, vals['count'], vals['u']])
                for c in range(1, len(headers3) + 1):
                    ws3.cell(row=row_idx, column=c).border = border
                row_idx += 1
        autosize(ws3)
        ws3.freeze_panes = 'A2'

        # Return the file
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="rack_utilization_report.xlsx"'
        return response


class RackReportPDFExportView(LoginRequiredMixin, View):

    def get(self, request):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return HttpResponse('reportlab is not installed. Run: pip install reportlab', status=500)

        results = compute_utilization()
        styles = getSampleStyleSheet()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
        )

        elements = []
        elements.append(Paragraph('Rack Utilization Report', styles['Title']))
        elements.append(Spacer(1, 0.5 * cm))

        table_data = [['Site', 'Location', 'Racks', 'Devices', 'Total U', 'Used U', 'Free U', '% Utilization', 'Alert']]
        row_alert_flags = [False]  # header row, no alert

        for r in results:
            table_data.append([
                r['site'].name,
                r['location'].name if r['location'] else 'No location',
                str(r['rack_count']),
                str(r['device_count']),
                str(r['total_u']),
                str(r['used_u']),
                str(r['free_u']),
                f"{r['pct']}%",
                f"> {ALERT_THRESHOLD}%" if r['alert'] else 'OK',
            ])
            row_alert_flags.append(r['alert'])

        # Totals row
        total_u_sum = sum(r['total_u'] for r in results)
        used_u_sum = sum(r['used_u'] for r in results)
        table_data.append([
            'TOTAL', '',
            str(sum(r['rack_count'] for r in results)),
            str(sum(r['device_count'] for r in results)),
            str(total_u_sum),
            str(used_u_sum),
            str(total_u_sum - used_u_sum),
            f"{round(used_u_sum / total_u_sum * 100, 1) if total_u_sum else 0}%",
            '',
        ])
        row_alert_flags.append(False)

        col_widths = [4.5 * cm, 4 * cm, 2 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 3 * cm, 2.5 * cm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E0E0E0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]

        # Highlight alert rows in red
        for idx, is_alert in enumerate(row_alert_flags):
            if is_alert:
                style_commands.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#FFCDD2')))

        table.setStyle(TableStyle(style_commands))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rack_utilization_report.pdf"'
        return response
